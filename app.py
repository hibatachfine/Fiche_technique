import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ======================
# Authentification
# ======================
def check_password():
    def password_entered():
        if st.session_state.get("password") == "FT.petitforestier":
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("Mot de passe", type="password", on_change=password_entered, key="password")
        st.stop()
    elif not st.session_state["password_correct"]:
        st.text_input("Mot de passe", type="password", on_change=password_entered, key="password")
        st.error("Mot de passe incorrect")
        st.stop()

check_password()

# ======================
# Interface
# ======================
st.image("petit_forestier_logo_officiel.png", width=700)
st.markdown("<h1 style='color:#057A20;'>Générateur de Fiches Techniques</h1>", unsafe_allow_html=True)
st.markdown("---")

# ======================
# Chargement des données
# ======================
try:
    df = pd.read_excel("bdd_ht.xlsx", sheet_name="FS_referentiel_produits_std")
    cabine_df = pd.read_excel("bdd_ht.xlsx", sheet_name="CABINES")
    chassis_df = pd.read_excel("bdd_ht.xlsx", sheet_name="CHASSIS")
    caisse_df  = pd.read_excel("bdd_ht.xlsx", sheet_name="CAISSES")
    moteur_df  = pd.read_excel("bdd_ht.xlsx", sheet_name="MOTEURS")
    frigo_df   = pd.read_excel("bdd_ht.xlsx", sheet_name="FRIGO")
    hayon_df   = pd.read_excel("bdd_ht.xlsx", sheet_name="HAYONS")
except Exception as e:
    st.error(f"Erreur lors du chargement des fichiers : {e}")
    st.stop()

# Normalisation colonnes
df.columns         = df.columns.str.replace('\n', ' ').str.strip()
cabine_df.columns  = cabine_df.columns.str.strip()
chassis_df.columns = chassis_df.columns.str.strip()
caisse_df.columns  = caisse_df.columns.str.strip()
moteur_df.columns  = moteur_df.columns.str.strip()
frigo_df.columns   = frigo_df.columns.str.strip()
hayon_df.columns   = hayon_df.columns.str.strip()

# ======================
# Filtres
# ======================
code_pays = st.selectbox("Code pays", sorted(df["Code_Pays"].dropna().unique()))
df_filtered = df[df["Code_Pays"] == code_pays]

marque = st.selectbox("Marque", sorted(df_filtered["Marque"].dropna().unique()))
df_filtered = df_filtered[df_filtered["Marque"] == marque]

modele = st.selectbox("Modèle", sorted(df_filtered["Modele"].dropna().unique()))
df_filtered = df_filtered[df_filtered["Modele"] == modele]

code_pf = st.selectbox("Code PF", sorted(df_filtered["Code_PF"].dropna().unique()))
df_filtered = df_filtered[df_filtered["Code_PF"] == code_pf]

# Standard PF (facultatif)
if "Standard_PF" in df_filtered.columns and not df_filtered["Standard_PF"].dropna().empty:
    standard_pf = st.selectbox("Standard PF", sorted(df_filtered["Standard_PF"].dropna().unique()))
    df_filtered = df_filtered[df_filtered["Standard_PF"] == standard_pf]
else:
    standard_pf = ""

code_cabine = st.selectbox("Cabine", df_filtered["C_Cabine"].dropna().unique())
code_chassis = st.selectbox("Châssis", df_filtered["C_Chassis"].dropna().unique())
code_caisse  = st.selectbox("Caisse",  df_filtered["C_Caisse"].dropna().unique())
code_moteur  = st.selectbox("Moteur",  df_filtered["M_Moteur"].dropna().unique())
code_frigo   = st.selectbox("Groupe Frigorifique", df_filtered["C_Groupe Frigorifique"].dropna().unique())
code_hayon   = st.selectbox("Hayon",   df_filtered["C_Hayon"].dropna().unique())

# ======================
# Utilitaires
# ======================
def to_cell_value(x):
    if pd.isna(x):
        return ""
    return x

def safe_write(ws, cell_ref, value):
    value = to_cell_value(value)
    col_letters = ''.join(filter(str.isalpha, cell_ref))
    row_number  = int(''.join(filter(str.isdigit, cell_ref)))
    col_index   = column_index_from_string(col_letters)

    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            min_col, min_row, *_ = merged.bounds
            col_index = min_col
            row_number = min_row
            break

    ws.cell(row=row_number, column=col_index).value = value

def get_criteria_list(df_comp, code, code_column):
    row = df_comp[df_comp[code_column] == code]
    if row.empty:
        return []
    row = row.iloc[0].dropna()
    exclude = {code_column, 'Produit (P) / Option (O)'}
    out = []
    for col, val in row.items():
        if col in exclude:
            continue
        s = str(val).strip()
        if s and s.lower() != "nan":
            out.append(s)
    return out[1:] if len(out) > 1 else []  # on saute la 1ère valeur

def insert_criteria(ws, start_cell, criteria_list):
    col = ''.join(filter(str.isalpha, start_cell))
    row0 = int(''.join(filter(str.isdigit, start_cell)))
    for i, item in enumerate(criteria_list):
        safe_write(ws, f"{col}{row0 + i}", item)

def insert_criteria_extended(ws, start_cell, criteria_list, overflow_col="D", max_rows=7):
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))

    for i, item in enumerate(criteria_list):
        if i < max_rows:
            col = start_col
            row = start_row + i
        else:
            col = overflow_col
            row = start_row + (i - max_rows)

        cell_ref = f"{col}{row}"
        try:
            ws[cell_ref] = str(item).strip()
        except Exception as e:
            print(f"Erreur cellule {cell_ref} : {e}")

# ======================
# Génération FT
# ======================
def generate_filled_ft():
    wb = load_workbook("Modèle FT.xlsx")
    ws = wb["TYPE_FROID"]

    sel = df[
        (df["Code_Pays"] == code_pays) &
        (df["Marque"]    == marque) &
        (df["Modele"]    == modele) &
        (df["Code_PF"]   == code_pf)
    ]
    if standard_pf:
        sel = sel[sel["Standard_PF"] == standard_pf]

    if sel.empty:
        sel = df[df["Code_PF"] == code_pf]

    selected_row = sel.iloc[0]

    # Dimensions
    safe_write(ws, "H6",  selected_row.get("W int  utile  sur plinthe", ""))
    safe_write(ws, "H7",  selected_row.get("L int  utile  sur plinthe", ""))
    safe_write(ws, "H8",  selected_row.get("H int", ""))
    safe_write(ws, "H9",  selected_row.get("H", ""))  
    safe_write(ws, "J5",  selected_row.get("L", ""))
    safe_write(ws, "J6",  selected_row.get("Z", ""))
    safe_write(ws, "J7",  selected_row.get("Hc", ""))
    safe_write(ws, "J8",  selected_row.get("F", ""))
    safe_write(ws, "J9",  selected_row.get("X", ""))

    # Bloc PTAC
    safe_write(ws, "G11", selected_row.get("PTAC", ""))
    safe_write(ws, "G12", selected_row.get("CU", ""))
    safe_write(ws, "G13", selected_row.get("Volume", ""))
    safe_write(ws, "G14", selected_row.get("palettes 800 x 1200 mm", ""))

    # En-tête
    entete = f"{marque}     {modele}     {code_pf}     {standard_pf}"
    safe_write(ws, "B1", entete)

    # Composants
    insert_criteria(ws, "B18", get_criteria_list(cabine_df, code_cabine, "C_Cabine"))
    insert_criteria(ws, "D18", get_criteria_list(moteur_df, code_moteur, "M_Moteur"))
    insert_criteria(ws, "F18", get_criteria_list(chassis_df, code_chassis, "C_Chassis"))
    insert_criteria(ws, "B37", get_criteria_list(caisse_df,  code_caisse,  "C_Caisse"))
    insert_criteria_extended(ws, "B58", get_criteria_list(frigo_df, code_frigo, "C_Groupe Frigorifique"))
    insert_criteria_extended(ws, "B67", get_criteria_list(hayon_df, code_hayon, "C_Hayon"))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ======================
# Téléchargement
# ======================
st.download_button(
    label="Télécharger la fiche technique",
    data=generate_filled_ft(),
    file_name=f"FT_{code_pf}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
